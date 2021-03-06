import openpyxl
from . import utilits
from .constants import *
from .error_messages import *
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook
from typing import Union
from django.shortcuts import get_object_or_404
from ..models import Distributor

"""
ПАРСЕРЫ ДАННЫХ
"""


class DataParserOneFileCabinet:
    """Парсер данных для одного файла, отчет кабинет 007"""

    def __init__(self, file, bonus_count: int, sales_units: str) -> None:
        self.bonus_count: int = bonus_count
        self.sales_units: str = sales_units
        self.work_book: Workbook = openpyxl.load_workbook(file)
        self.work_sheet: Worksheet = self.work_book.get_sheet_by_name(self.work_book.get_sheet_names()[0])
        self.data: dict = self.get_managers_data()
        self.managers_data: list = self.data.get('managers_data')
        self.errors: list = self.data.get('errors')

    def get_managers_data(self) -> dict:
        iteration, free_cell = 0, 0
        managers_data, manager, manager_data, errors = [], [], [], []
        new_phone_number, new_manager_name = True, True
        while free_cell < 2:
            iteration += 1
            cell_a = self.work_sheet[f'A{iteration}']
            cell_b = self.work_sheet[f'B{iteration}']
            cell_c = self.work_sheet[f'C{iteration}']

            if cell_a.value and not cell_c.value:
                free_cell = 0

                if new_phone_number:
                    if utilits.validate_phone_number(cell_a.value):
                        manager.append(cell_a.value)
                        new_phone_number = False
                        continue
                    errors.append([cell_a.row, cell_a.value, cell_a.coordinate, PHONE_NUMBER_VALIDATION_ERROR])

                if not new_phone_number and new_manager_name:
                    if utilits.validate_manager_name(cell_a.value):
                        manager.append(cell_a.value)
                        new_manager_name = False
                        continue
                    errors.append([cell_a.row, cell_a.value, cell_a.coordinate, MANAGER_NAME_VALIDATION_ERROR])

                if not new_phone_number and not new_manager_name:
                    errors.append([cell_c.row, cell_c.value, cell_c.coordinate, CELL_VALUE_ERROR])
                    continue

            if cell_a.value and cell_c.value:
                if validation_result := utilits.validate_cell_value(cell_a, cell_c).get('error'):
                    errors.append(validation_result)
                    continue
                manager_data.append([cell_a.value, cell_b.value, cell_c.value])
                continue

            if not cell_a.value and not cell_b.value and not cell_c.value:
                manager.append(manager_data)
                managers_data.append(manager)
                manager, manager_data = [], []
                new_phone_number, new_manager_name = True, True
                free_cell += 1
                continue
        return {'managers_data': managers_data, 'errors': errors}


class DataParserTwoFileCabinet:
    """Парсер данных для двух файлов, отчет кабинет 007"""

    def __init__(self, file1, file2, bonus_count: int, distributor_name_input: str, sales_units: str) -> None:
        self.bonus_count: int = bonus_count
        self.sales_units: str = sales_units
        self.work_book_file1 = openpyxl.load_workbook(file1)
        self.work_sheet_file1 = self.work_book_file1.get_sheet_by_name(self.work_book_file1.get_sheet_names()[0])
        self.work_book_file2 = openpyxl.load_workbook(file2)
        self.work_sheet_file2 = self.work_book_file2.get_sheet_by_name(self.work_book_file2.get_sheet_names()[0])
        self.distributor = get_object_or_404(Distributor, name=distributor_name_input)
        self.data: dict = self.get_managers_data()
        self.managers_data: list = self.data.get('managers_data')
        self.errors: list = self.data.get('errors')

    def get_file1_data(self) -> list:
        iteration: int = 0
        file1_data: list = []
        while True:
            iteration += 1
            cell_a = self.work_sheet_file1[f'A{iteration}']
            cell_b = self.work_sheet_file1[f'B{iteration}']
            if not cell_a.value and not cell_b.value:
                break
            if cell_a.value and not cell_b.value:
                file1_data.append({'manager_name': cell_a, 'data': []})
                continue
            if cell_a.value and cell_b.value:
                file1_data[-1]['data'].append((cell_a, cell_b))
                continue
        return file1_data

    def get_file2_data(self) -> list:
        iteration: int = 0
        file2_data: list = []
        while True:
            iteration += 1
            cell_a = self.work_sheet_file2[f'A{iteration}']
            cell_b = self.work_sheet_file2[f'B{iteration}']
            cell_c = self.work_sheet_file2[f'C{iteration}']
            cell_d = self.work_sheet_file2[f'D{iteration}']
            if not cell_c.value and not cell_d.value:
                break
            if cell_c.value and cell_d.value:
                file2_data.append([cell_a, cell_b, cell_c, cell_d])
                continue
        return file2_data

    def get_managers_data(self) -> dict:
        file1_data: list = self.get_file1_data()
        file2_data: list = self.get_file2_data()
        external_id: str = self.distributor.external_id
        managers_data, manager, manager_data, errors = [], [], [], []
        for data in file1_data:
            manager_name = data.get('manager_name')
            if phone_number := utilits.search_phone_number(external_id, manager_name.value, file2_data):

                if utilits.validate_phone_number(phone_number):
                    manager.append(phone_number)
                else:
                    errors.append(
                        [manager_name.row, phone_number, manager_name.coordinate, PHONE_NUMBER_VALIDATION_ERROR])
                    continue

                if utilits.validate_manager_name(manager_name.value):
                    manager.append(manager_name.value)
                else:
                    errors.append(
                        [manager_name.row, manager_name, manager_name.coordinate, MANAGER_NAME_VALIDATION_ERROR])
                    continue

                for tpl in data.get('data'):
                    nomenclature_cell_object = tpl[0]
                    mass_cell_object = tpl[1]
                    if validation_result := utilits.validate_cell_value(nomenclature_cell_object, mass_cell_object).get(
                            'errors'):
                        errors.append(validation_result)
                        continue
                    manager_data.append([nomenclature_cell_object.value, mass_cell_object.value])
                    continue
                manager.append(manager_data)
                managers_data.append(manager)
                manager, manager_data = [], []
            else:
                errors.append(
                    [manager_name.row, manager_name.value, manager_name.coordinate, TELEPHONE_NUMBER_SEARCH_ERROR])
                continue
        managers_data.append([[]])
        return {'managers_data': managers_data, 'errors': errors}


class DataParserOneFileNotManagerCabinet:
    """Парсер данных для одного файла, отчет без менеджеров

       ЭТОТ КЛАСС РАБОТАЕТ ПО УПРОЩЕННОЙ СХЕМЕ ИЗВЛЕЧЕНИЯ ДАННЫХ!!!
       СТРУКТУРА ДАННЫХ ПАРСЕРА И МЕТОДЫ ИХ ПОЛУЧЕНИЯ ОТЛИЧАЮТСЯ ОТ ПРЕДЫДУЩИХ КЛАССОВ-ПАРСЕРОВ!!!
       НЕ НАСЛЕДОВАТЬСЯ ОТ ДАННОГО КЛАССА ПРЕДВАРИТЕЛЬНО НЕ УБЕДИВШИСЬ В ВАЛИДНОСТИ НАСЛЕДОВАНИЯ!!!
    """

    def __init__(self, file, bonus_count: int, sales_units: str) -> None:
        self.bonus_count: int = bonus_count
        self.sales_units: str = sales_units
        self.work_book: Workbook = openpyxl.load_workbook(file)
        self.work_sheet: Worksheet = self.work_book.get_sheet_by_name(self.work_book.get_sheet_names()[0])
        self.data: dict = self.parse_data()
        self.data_list: list = self.data.get('data_list')
        self.errors: list = self.data.get('errors')

    def parse_data(self) -> dict:
        iteration: int = 0
        data_list, errors = [], []
        while True:
            iteration += 1
            nomenclature_cell = self.work_sheet[f'A{iteration}']
            mass_cell = self.work_sheet[f'B{iteration}']

            if not nomenclature_cell.value and not mass_cell.value:
                break
            elif valid_result := utilits.validate_cell_value(nomenclature_cell, mass_cell).get('errors'):
                errors.append(valid_result)
                continue
            else:
                data_list.append([nomenclature_cell.value, mass_cell.value])
                continue
        return {'data_list': data_list, 'errors': errors}


"""ОБРАБОТЧИКИ ДАННЫХ"""


# ОБРАБОТЧИКИ ДЛЯ ФОРМА ОТЧЕТА: КАБИНЕТ 007


class CaseCabinetBugBonus(DataParserOneFileCabinet):
    """Обработчик данных для кейса: кабинет 007, бонус за мешок"""

    def __init__(self, file, bonus_count: int, sales_units: str) -> None:
        super().__init__(file, bonus_count, sales_units)
        self.report_file: Workbook = openpyxl.Workbook()
        self.work_report: Worksheet = self.report_file.create_sheet('Рабочий отчет', 0)
        self.cabinet007_report: Worksheet = self.report_file.create_sheet('Отчет для сдачи', 1)

    def get_cabinet_report(self) -> None:
        iteration: int = 1
        for manager in self.managers_data:
            total_bonus_for_this_manager: int = 0
            if not self.set_telephone_and_name_in_cabinet_report(iteration, manager):
                break
            iteration += 2

            for manager_data in manager[2]:
                total_bonus_for_this_manager += self.get_bonus_sum(manager_data)

            self.cabinet007_report[f'B{iteration}'] = total_bonus_for_this_manager / DUMMY_BONUS_PER_TON
            iteration += 2

    def set_telephone_and_name_in_cabinet_report(self, iteration: int, manager: list) -> bool:
        if telephone_number := manager[0]:
            self.cabinet007_report[f'A{iteration}'] = telephone_number
            iteration += 1
            manager_name = manager[1]
            self.cabinet007_report[f'A{iteration}'] = manager_name
            iteration += 1
            self.cabinet007_report[f'A{iteration}'] = '00208'
            return True
        return False

    def get_bonus_sum(self, manager_data: list) -> float:
        nomenclature: str = manager_data[0]
        weight: float = float(manager_data[2])
        product_mass: int = utilits.get_product_mass(nomenclature)

        if self.sales_units == 'tons':
            bags_count = weight * COUNT_KGS_IN_TON / product_mass
        else:
            bags_count = weight

        bonus_sum = bags_count * self.bonus_count
        return bonus_sum

    def get_work_report(self) -> None:
        self.set_work_report_title()
        iteration = 1
        total_bags_count, total_bonus_sum = 0, 0
        for manager in self.managers_data:
            if not self.set_telephone_and_name_in_work_report(iteration, manager):
                break
            iteration += 3
            for manager_data in manager[2]:
                calculation = self.get_calculations_for_manager_data(manager_data)
                total_bags_count += calculation['bags_count']
                total_bonus_sum += calculation['bonus_sum']
                self.set_work_report_cell_value(iteration, calculation)
                iteration += 1
        self.summarize(iteration, total_bags_count, total_bonus_sum)

    def summarize(self, iteration: int, total_bags_count: int, total_bonus_sum: int) -> None:
        self.work_report[f'A{iteration}'] = 'Итого:'
        self.work_report[f'E{iteration}'] = total_bags_count
        self.work_report[f'G{iteration}'] = total_bonus_sum

    def set_telephone_and_name_in_work_report(self, iteration: int, manager: list) -> bool:
        if telephone_number := manager[0]:
            manager_name = manager[1]
            iteration += 1
            self.work_report[f'A{iteration}'] = telephone_number
            iteration += 1
            self.work_report[f'A{iteration}'] = manager_name
            return True
        return False

    def set_work_report_title(self) -> None:
        self.work_report['B1'] = 'Номенклатурный код'
        self.work_report['C1'] = 'Тоннаж'
        self.work_report['D1'] = 'Вес единицы продукта'
        self.work_report['E1'] = 'Количество мешков'
        self.work_report['F1'] = 'Бонус за 1 мешок'
        self.work_report['G1'] = 'Сумма бонуса'

    def set_work_report_cell_value(self, iteration: int, calculation: dict) -> None:
        self.work_report[f'A{iteration}'] = calculation['nomenclature']
        self.work_report[f'B{iteration}'] = calculation['nomenclature_code']
        self.work_report[f'C{iteration}'] = calculation['weight']
        self.work_report[f'D{iteration}'] = calculation['product_mass']
        self.work_report[f'E{iteration}'] = calculation['bags_count']
        self.work_report[f'F{iteration}'] = calculation['bonus_count']
        self.work_report[f'G{iteration}'] = calculation['bonus_sum']

    def get_calculations_for_manager_data(self, manager_data: list) -> dict:
        nomenclature: str = manager_data[0]
        nomenclature_code: str = manager_data[1]
        weight: float = float(manager_data[2])
        product_mass: int = utilits.get_product_mass(nomenclature)

        if self.sales_units == 'tons':
            bags_count = weight * COUNT_KGS_IN_TON / product_mass
        else:
            bags_count = weight

        bonus_count = self.bonus_count
        bonus_sum = self.get_bonus_sum(manager_data)
        return {'nomenclature': nomenclature, 'nomenclature_code': nomenclature_code,
                'weight': weight, 'product_mass': product_mass, 'bags_count': bags_count,
                'bonus_count': bonus_count, 'bonus_sum': bonus_sum}


class CaseCabinetFixedBonusPalette(CaseCabinetBugBonus):
    """Обработчик данных для кейса: кабинет 007, фиксированный бонус с палетты"""

    def __init__(self, file, bonus_count: int, product_count_input: int,
                 action_checkbox: bool, sales_units: str) -> None:
        super().__init__(file, bonus_count, sales_units)
        self.product_count_input: int = product_count_input
        self.action: bool = action_checkbox

    def get_bonus_sum(self, manager_data: list) -> int:
        nomenclature: str = manager_data[0]
        weight: float = float(manager_data[2])
        product_mass: int = utilits.get_product_mass(nomenclature)

        if self.sales_units == 'tons':
            bags_count = weight * COUNT_KGS_IN_TON / product_mass
        else:
            bags_count = weight

        palette_count = bags_count // BUGS_COUNT_IN_PALETTE[str(product_mass)]
        if palette_count < self.product_count_input:
            return 0
        if self.action:
            return palette_count // self.product_count_input * self.bonus_count
        else:
            return self.bonus_count


class CaseCabinetFixedBonusBugs(CaseCabinetBugBonus):
    """Обработчик данных для кейса: кабинет 007, фиксированный бонус с мешка"""

    def __init__(self, file, bonus_count: int, product_count_input: int,
                 action_checkbox: bool, sales_units: str) -> None:
        super().__init__(file, bonus_count, sales_units)
        self.product_count_input: int = product_count_input
        self.action: bool = action_checkbox

    def get_bonus_sum(self, manager_data: list) -> Union[int, float]:
        nomenclature: str = manager_data[0]
        weight: float = float(manager_data[2])
        product_mass: int = utilits.get_product_mass(nomenclature)

        if self.sales_units == 'tons':
            bags_count = weight * COUNT_KGS_IN_TON / product_mass
        else:
            bags_count = weight

        if bags_count < self.product_count_input:
            return 0
        if self.action:
            return bags_count // self.product_count_input * self.bonus_count
        else:
            return self.bonus_count


class CaseCabinetFixedBonusTons(CaseCabinetBugBonus):
    """Обработчик данных для кейса: кабинет 007, фиксированный бонус с тонны"""

    def __init__(self, file, bonus_count: int, product_count_input: int,
                 action_checkbox: bool, sales_units: str) -> None:
        super().__init__(file, bonus_count, sales_units)
        self.product_count_input: int = product_count_input
        self.action: bool = action_checkbox

    def get_bonus_sum(self, manager_data: list) -> Union[int, float]:
        nomenclature: str = manager_data[0]
        weight: float = float(manager_data[2])
        product_mass: int = utilits.get_product_mass(nomenclature)

        if self.sales_units == 'tons':
            bags_count = weight
        else:
            bags_count = weight * product_mass

        if bags_count < self.product_count_input:
            return 0
        if self.action:
            return bags_count // self.product_count_input * self.bonus_count
        else:
            return self.bonus_count


# ОБРАБОТЧИКИ ДЛЯ ФОРМА ОТЧЕТА: ОТЧЕТ С МЕНЕДЖЕРАМИ


class CaseManagersBugBonus(DataParserTwoFileCabinet):
    """Обработчик данных для кейса: отчет с менеджерами, бонус за мешок"""

    def __init__(self, file1, file2, bonus_count: int, distributor_name_input: str, sales_units: str) -> None:
        super().__init__(file1, file2, bonus_count, distributor_name_input, sales_units)
        self.report_file: Workbook = openpyxl.Workbook()
        self.work_report: Worksheet = self.report_file.create_sheet('Рабочий отчет', 0)
        self.cabinet007_report: Worksheet = self.report_file.create_sheet('Отчет для сдачи', 1)

    def get_cabinet_report(self) -> None:
        iteration: int = 1
        for manager in self.managers_data:
            total_bonus_for_this_manager: int = 0
            if not self.set_telephone_and_name_in_cabinet_report(iteration, manager):
                break
            iteration += 2

            for manager_data in manager[2]:
                total_bonus_for_this_manager += self.get_bonus_sum(manager_data)

            self.cabinet007_report[f'B{iteration}'] = total_bonus_for_this_manager / DUMMY_BONUS_PER_TON
            iteration += 2

    def set_telephone_and_name_in_cabinet_report(self, iteration: int, manager: list) -> bool:
        if telephone_number := manager[0]:
            self.cabinet007_report[f'A{iteration}'] = telephone_number
            iteration += 1
            manager_name = manager[1]
            self.cabinet007_report[f'A{iteration}'] = manager_name
            iteration += 1
            self.cabinet007_report[f'A{iteration}'] = '00208'
            return True
        return False

    def get_work_report(self) -> None:
        iteration: int = 1
        total_bags_count, total_bonus_sum = 0, 0
        self.set_work_report_title()
        for manager in self.managers_data:
            if not self.set_telephone_and_name_in_work_report(iteration, manager):
                break
            iteration += 2
            for manager_data in manager[2]:
                iteration += 1
                calculation = self.get_calculations_for_manager_data(manager_data)
                total_bags_count += calculation['bags_count']
                total_bonus_sum += calculation['bonus_sum']
                self.set_work_report_cell_value(iteration, calculation)
        self.summarize(iteration, total_bags_count, total_bonus_sum)

    def summarize(self, iteration: int, total_bags_count: int, total_bonus_sum: int) -> None:
        iteration += 1
        self.work_report[f'A{iteration}'] = 'Итого:'
        self.work_report[f'D{iteration}'] = total_bags_count
        self.work_report[f'F{iteration}'] = total_bonus_sum

    def set_telephone_and_name_in_work_report(self, iteration: int, manager: list) -> bool:
        if telephone_number := manager[0]:
            manager_name = manager[1]
            iteration += 1
            self.work_report[f'A{iteration}'] = telephone_number
            iteration += 1
            self.work_report[f'A{iteration}'] = manager_name
            return True
        return False

    def set_work_report_title(self) -> None:
        self.work_report['C1'] = 'Вес единицы продукта'
        self.work_report['D1'] = 'Количество мешков'
        self.work_report['E1'] = 'Бонус за 1 мешок'
        self.work_report['F1'] = 'Сумма бонуса'

    def set_work_report_cell_value(self, iteration: int, calculation: dict) -> None:
        self.work_report[f'A{iteration}'] = calculation['nomenclature']
        self.work_report[f'B{iteration}'] = calculation['weight']
        self.work_report[f'C{iteration}'] = calculation['product_mass']
        self.work_report[f'D{iteration}'] = calculation['bags_count']
        self.work_report[f'E{iteration}'] = calculation['bonus_count']
        self.work_report[f'F{iteration}'] = calculation['bonus_sum']

    def get_calculations_for_manager_data(self, manager_data: list) -> dict:
        nomenclature: str = manager_data[0]
        weight: float = float(manager_data[1])
        product_mass: int = utilits.get_product_mass(nomenclature)

        if self.sales_units == 'tons':
            bags_count = weight * COUNT_KGS_IN_TON / product_mass
        else:
            bags_count = weight

        bonus_count = self.bonus_count
        bonus_sum = self.get_bonus_sum(manager_data)
        return {'nomenclature': nomenclature, 'weight': weight, 'product_mass': product_mass,
                'bags_count': bags_count, 'bonus_count': bonus_count, 'bonus_sum': bonus_sum}

    def get_bonus_sum(self, manager_data: list) -> float:
        nomenclature: str = manager_data[0]
        weight: float = float(manager_data[1])
        product_mass: int = utilits.get_product_mass(nomenclature)

        if self.sales_units == 'tons':
            bags_count = weight * COUNT_KGS_IN_TON / product_mass
        else:
            bags_count = weight

        bonus_sum = bags_count * self.bonus_count
        return bonus_sum


class CaseManagersFixedBonusPalette(CaseManagersBugBonus):
    """Обработчик данных для кейса: отчет с менеджерами, фиксированный бонус с палетты"""

    def __init__(self, file1, file2, bonus_count: int, distributor_name_input: str,
                 sales_units: str, product_count_input: int, action_checkbox: bool) -> None:
        super().__init__(file1, file2, bonus_count, distributor_name_input, sales_units)
        self.product_count_input: int = product_count_input
        self.action: bool = action_checkbox

    def get_bonus_sum(self, manager_data: list) -> int:
        nomenclature: str = manager_data[0]
        weight: float = float(manager_data[2])
        product_mass: int = utilits.get_product_mass(nomenclature)

        if self.sales_units == 'tons':
            bags_count = weight * COUNT_KGS_IN_TON / product_mass
        else:
            bags_count = weight

        palette_count = bags_count // BUGS_COUNT_IN_PALETTE[str(product_mass)]
        if palette_count < self.product_count_input:
            return 0
        if self.action:
            return palette_count // self.product_count_input * self.bonus_count
        else:
            return self.bonus_count


class CaseManagersFixedBonusBugs(CaseManagersBugBonus):
    """Обработчик данных для кейса: отчет с менеджерами, фиксированный бонус с мешка"""

    def __init__(self, file1, file2, bonus_count: int, distributor_name_input: str,
                 sales_units: str, product_count_input: int, action_checkbox: bool) -> None:
        super().__init__(file1, file2, bonus_count, distributor_name_input, sales_units)
        self.product_count_input: int = product_count_input
        self.action: bool = action_checkbox

    def get_bonus_sum(self, manager_data: list) -> Union[int, float]:
        nomenclature: str = manager_data[0]
        weight: float = float(manager_data[2])
        product_mass: int = utilits.get_product_mass(nomenclature)

        if self.sales_units == 'tons':
            bags_count = weight * COUNT_KGS_IN_TON / product_mass
        else:
            bags_count = weight

        if bags_count < self.product_count_input:
            return 0
        if self.action:
            return bags_count // self.product_count_input * self.bonus_count
        else:
            return self.bonus_count


class CaseManagersFixedBonusTons(CaseManagersBugBonus):
    """Обработчик данных для кейса: отчет с менеджерами, фиксированный бонус с тонны"""

    def __init__(self, file1, file2, bonus_count: int, distributor_name_input: str,
                 sales_units: str, product_count_input: int, action_checkbox: bool) -> None:
        super().__init__(file1, file2, bonus_count, distributor_name_input, sales_units)
        self.product_count_input: int = product_count_input
        self.action: bool = action_checkbox

    def get_bonus_sum(self, manager_data: list) -> Union[int, float]:
        nomenclature: str = manager_data[0]
        weight: float = float(manager_data[2])
        product_mass: int = utilits.get_product_mass(nomenclature)

        if self.sales_units == 'tons':
            bags_count = weight
        else:
            bags_count = weight * product_mass

        if bags_count < self.product_count_input:
            return 0
        if self.action:
            return bags_count // self.product_count_input * self.bonus_count
        else:
            return self.bonus_count


# ОБРАБОТЧИКИ ДЛЯ ФОРМА ОТЧЕТА: ОТЧЕТ БЕЗ МЕНЕДЖЕРОВ


class CaseNotManagersBugBonus(DataParserOneFileNotManagerCabinet):
    """Обработчик данных для кейса: отчет без менеджеров, бонус за мешок"""

    def __init__(self, file, bonus_count: int, sales_units: str) -> None:
        super().__init__(file, bonus_count, sales_units)
        self.report_file: Workbook = openpyxl.Workbook()
        self.work_report: Worksheet = self.report_file.create_sheet('Рабочий отчет', 0)
        self.cabinet007_report: Worksheet = self.report_file.create_sheet('Отчет для сдачи', 1)

    def get_cabinet_report(self) -> None:
        total_bonus_for_this_manager: int = 0
        for data in self.data_list:
            total_bonus_for_this_manager += self.get_bonus_sum(data)
        self.cabinet007_report['A1'] = '00208'
        self.cabinet007_report['B1'] = total_bonus_for_this_manager / DUMMY_BONUS_PER_TON

    def get_work_report(self) -> None:
        iteration: int = 1
        total_bags_count, total_bonus_sum = 0, 0
        self.set_work_report_title()
        for data in self.data_list:
            iteration += 1
            calculation = self.get_calculations(data)
            total_bags_count += calculation['bags_count']
            total_bonus_sum += calculation['bonus_sum']
            self.set_work_report_cell_value(iteration, calculation)
        self.summarize(iteration, total_bags_count, total_bonus_sum)

    def set_work_report_title(self) -> None:
        self.work_report['C1'] = 'Вес единицы продукта'
        self.work_report['D1'] = 'Количество мешков'
        self.work_report['E1'] = 'Бонус за 1 мешок'
        self.work_report['F1'] = 'Сумма бонуса'

    def summarize(self, iteration: int, total_bags_count: int, total_bonus_sum: int) -> None:
        self.work_report[f'A{iteration}'] = 'Итого:'
        self.work_report[f'D{iteration}'] = total_bags_count
        self.work_report[f'F{iteration}'] = total_bonus_sum

    def get_calculations(self, data: list) -> dict:
        nomenclature: str = data[0]
        weight: float = float(data[1])
        product_mass: int = utilits.get_product_mass(nomenclature)

        if self.sales_units == 'tons':
            bags_count = weight * COUNT_KGS_IN_TON / product_mass
        else:
            bags_count = weight

        bonus_count: int = self.bonus_count
        bonus_sum: float = self.get_bonus_sum(data)
        return {'nomenclature': nomenclature, 'weight': weight, 'product_mass': product_mass,
                'bags_count': bags_count, 'bonus_count': bonus_count, 'bonus_sum': bonus_sum}

    def get_bonus_sum(self, data: list) -> float:
        nomenclature: str = data[0]
        weight: float = float(data[1])
        product_mass: int = utilits.get_product_mass(nomenclature)

        if self.sales_units == 'tons':
            bags_count = weight * COUNT_KGS_IN_TON / product_mass
        else:
            bags_count = weight

        bonus_sum = bags_count * self.bonus_count
        return bonus_sum

    def set_work_report_cell_value(self, iteration: int, calculation: dict) -> None:
        self.work_report[f'A{iteration}'] = calculation['nomenclature']
        self.work_report[f'B{iteration}'] = calculation['weight']
        self.work_report[f'C{iteration}'] = calculation['product_mass']
        self.work_report[f'D{iteration}'] = calculation['bags_count']
        self.work_report[f'E{iteration}'] = calculation['bonus_count']
        self.work_report[f'F{iteration}'] = calculation['bonus_sum']


class CaseNotManagersFixedBonusPalette(CaseNotManagersBugBonus):
    """Обработчик данных для кейса: отчет без менеджеров, фиксированный бонус с палетты"""

    def __init__(self, file, bonus_count: int, product_count_input: int,
                 action_checkbox: bool, sales_units: str) -> None:
        super().__init__(file, bonus_count, sales_units)
        self.product_count_input: int = product_count_input
        self.action: bool = action_checkbox

    def get_bonus_sum(self, data: list) -> float:
        nomenclature: str = data[0]
        weight: float = float(data[1])
        product_mass: int = utilits.get_product_mass(nomenclature)

        if self.sales_units == 'tons':
            bags_count = weight * COUNT_KGS_IN_TON / product_mass
        else:
            bags_count = weight

        palette_count = bags_count // BUGS_COUNT_IN_PALETTE[str(product_mass)]
        if palette_count < self.product_count_input:
            return 0
        if self.action:
            return palette_count // self.product_count_input * self.bonus_count
        else:
            return self.bonus_count


class CaseNotManagersFixedBonusBugs(CaseNotManagersBugBonus):
    """Обработчик данных для кейса: отчет без менеджеров, фиксированный бонус с мешка"""

    def __init__(self, file, bonus_count: int, product_count_input: int,
                 action_checkbox: bool, sales_units: str) -> None:
        super().__init__(file, bonus_count, sales_units)
        self.product_count_input: int = product_count_input
        self.action: bool = action_checkbox

    def get_bonus_sum(self, data: list) -> float:
        nomenclature: str = data[0]
        weight: float = float(data[1])
        product_mass: int = utilits.get_product_mass(nomenclature)

        if self.sales_units == 'tons':
            bags_count = weight * COUNT_KGS_IN_TON / product_mass
        else:
            bags_count = weight

        if bags_count < self.product_count_input:
            return 0
        if self.action:
            return bags_count // self.product_count_input * self.bonus_count
        else:
            return self.bonus_count


class CaseNotManagersFixedBonusTons(CaseNotManagersBugBonus):
    """Обработчик данных для кейса: отчет без менеджеров, фиксированный бонус с тонны"""

    def __init__(self, file, bonus_count: int, product_count_input: int,
                 action_checkbox: bool, sales_units: str) -> None:
        super().__init__(file, bonus_count, sales_units)
        self.product_count_input: int = product_count_input
        self.action: bool = action_checkbox

    def get_bonus_sum(self, data: list) -> float:
        nomenclature: str = data[0]
        weight: float = float(data[1])
        product_mass: int = utilits.get_product_mass(nomenclature)

        if self.sales_units == 'tons':
            bags_count = weight
        else:
            bags_count = weight * product_mass

        if bags_count < self.product_count_input:
            return 0
        if self.action:
            return bags_count // self.product_count_input * self.bonus_count
        else:
            return self.bonus_count
